"""Extra globe layers added 2026-05-23. Imported by layers.py and concatenated
to LAYERS so a `git restore` of layers.py only loses the import line, not all
the fetcher code in here.

Each fetcher returns the same {layer, updatedAt, count, items, error?} payload
shape via _payload(). Designed to be drop-in compatible with collector.py's
refresh() loop.
"""
from __future__ import annotations

import asyncio
import csv
import gzip
import io
import json
import math
import os
import sys
import time
import urllib.parse
import urllib.request
from datetime import datetime, timezone, timedelta
from pathlib import Path

# Shared fetcher (residential proxy, curl_cffi chrome impersonation).
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "_shared"))
from proxy_fetcher import ProxyFetcher, fetcher  # noqa: E402

_F = fetcher(impersonate="chrome146")


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _payload(layer: str, items: list, error: str | None = None) -> dict:
    p = {"layer": layer, "updatedAt": _now_iso(), "count": len(items), "items": items}
    if error:
        p["error"] = error
    return p


async def _aget_json(url: str, headers: dict | None = None, timeout: int = 25, tries: int = 2):
    """Async GET → JSON. Retries with fresh IP."""
    last_err = None
    for _ in range(tries):
        try:
            r = await _F.aget(url, headers=headers or {"Accept": "application/json"},
                              timeout=timeout)
            if r.status == 200 and r.body:
                return json.loads(r.body)
            last_err = RuntimeError(f"status={r.status}")
        except Exception as e:
            last_err = e
    raise last_err if last_err else RuntimeError("unknown")


def _read_latest_boats_blob() -> list:
    """Pull latest boats blob from Vercel Blob (collector token can read on
    same prefix). Returns the items list, [] on any failure."""
    try:
        tok = os.environ.get("BLOB_READ_WRITE_TOKEN", "")
        if not tok:
            return []
        req = urllib.request.Request(
            "https://blob.vercel-storage.com/?prefix=globe/boats.json&limit=1",
            headers={"Authorization": f"Bearer {tok}"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            blobs = json.load(resp).get("blobs") or []
        if not blobs:
            return []
        with urllib.request.urlopen(blobs[0]["url"], timeout=30) as resp:
            return (json.load(resp).get("items") or [])
    except Exception:
        return []


# ============================================================================
# Weather: tsunami warnings (NWS) + aurora (NOAA Ovation)
# ============================================================================
NWS_UA = "InspoGlobe/1.0 (+zac@zacstern.co)"
NWS_TSUNAMI_URL = (
    "https://api.weather.gov/alerts/active"
    "?event=Tsunami%20Warning&event=Tsunami%20Advisory&event=Tsunami%20Watch"
    "&status=actual"
)


def _polygon_centroid(coords):
    minLat = 90.0; maxLat = -90.0; minLng = 180.0; maxLng = -180.0; n = 0

    def visit(c):
        nonlocal minLat, maxLat, minLng, maxLng, n
        if isinstance(c, list) and len(c) >= 2 and isinstance(c[0], (int, float)):
            lng, lat = c[0], c[1]
            if lng < minLng: minLng = lng
            if lng > maxLng: maxLng = lng
            if lat < minLat: minLat = lat
            if lat > maxLat: maxLat = lat
            n += 1
        elif isinstance(c, list):
            for x in c:
                visit(x)

    visit(coords)
    if not n:
        return None, None
    return (minLat + maxLat) / 2.0, (minLng + maxLng) / 2.0


async def fetch_tsunami():
    try:
        data = await _aget_json(NWS_TSUNAMI_URL,
                                headers={"User-Agent": NWS_UA, "Accept": "application/geo+json"},
                                timeout=30, tries=2)
    except Exception:
        return _payload("tsunami", [])
    items = []
    for feat in (data.get("features") or []):
        props = feat.get("properties") or {}
        geom = feat.get("geometry") or {}
        gtype = geom.get("type")
        if gtype in ("Polygon", "MultiPolygon"):
            lat, lng = _polygon_centroid(geom.get("coordinates"))
        elif gtype == "Point":
            c = geom.get("coordinates") or []
            lng, lat = (c[0], c[1]) if len(c) >= 2 else (None, None)
        else:
            continue
        if lat is None:
            continue
        event = props.get("event") or "Tsunami"
        is_warning = "Warning" in event
        is_watch = "Watch" in event
        items.append({
            "id": props.get("id") or f"tsu-{len(items)}",
            "lat": lat, "lng": lng,
            "label": event,
            "headline": props.get("headline"),
            "severity": props.get("severity"),
            "effective": props.get("effective"),
            "expires": props.get("expires"),
            "areaDesc": props.get("areaDesc"),
            "category": event,
            "color": "#00B0FF" if is_warning else ("#FFC400" if is_watch else "#26C6DA"),
        })
    return _payload("tsunami", items)


OVATION_URL = "https://services.swpc.noaa.gov/json/ovation_aurora_latest.json"


def _aurora_color(prob):
    if prob >= 70: return "#FF3D3D"
    if prob >= 40: return "#FF9500"
    if prob >= 20: return "#FFEB3B"
    return "#76FF03"


async def fetch_aurora():
    try:
        data = await _aget_json(OVATION_URL, headers={"User-Agent": "globe-recon/1.0"},
                                timeout=20, tries=2)
    except Exception:
        return _payload("aurora", [])
    items = []
    candidates = []
    for c in (data.get("coordinates") or []):
        if len(c) < 3:
            continue
        prob = c[2]
        if prob is None or prob < 10:
            continue
        candidates.append(c)
    if len(candidates) > 500:
        candidates.sort(key=lambda c: -c[2])
        candidates = candidates[:500]
    forecast_time = data.get("Forecast Time")
    for i, c in enumerate(candidates):
        lng = c[0] - 360 if c[0] > 180 else c[0]
        items.append({
            "id": f"aurora-{i}",
            "lat": c[1], "lng": lng,
            "label": f"Aurora {c[2]}%",
            "probability": c[2],
            "forecast_time": forecast_time,
            "color": _aurora_color(c[2]),
        })
    return _payload("aurora", items)


# ============================================================================
# Tornado warnings (NWS) + Hurricanes (NHC)
# ============================================================================
NWS_TORNADO_URL = (
    "https://api.weather.gov/alerts/active"
    "?event=Tornado%20Warning&event=Tornado%20Watch"
    "&event=Severe%20Thunderstorm%20Warning&status=actual"
)


async def fetch_tornado_warnings_v2():
    try:
        data = await _aget_json(NWS_TORNADO_URL,
                                headers={"User-Agent": NWS_UA, "Accept": "application/geo+json"},
                                timeout=30, tries=2)
    except Exception:
        return _payload("tornado_warnings", [])
    items = []
    for feat in (data.get("features") or []):
        props = feat.get("properties") or {}
        geom = feat.get("geometry") or {}
        if geom.get("type") not in ("Polygon", "MultiPolygon"):
            continue
        lat, lng = _polygon_centroid(geom.get("coordinates"))
        if lat is None:
            continue
        event = props.get("event") or "Alert"
        items.append({
            "id": props.get("id") or f"nws-{len(items)}",
            "lat": lat, "lng": lng,
            "label": event,
            "headline": props.get("headline"),
            "severity": props.get("severity"),
            "effective": props.get("effective"),
            "expires": props.get("expires"),
            "areaDesc": props.get("areaDesc"),
            "category": event,
            "color": "#FF1744" if "Warning" in event else "#FFC400",
        })
    return _payload("tornado_warnings", items)


NHC_URL = "https://www.nhc.noaa.gov/CurrentStorms.json"


def _saffir_simpson(kts):
    try:
        v = int(kts)
    except (TypeError, ValueError):
        return None
    if v < 34: return "TD"
    if v < 64: return "TS"
    if v <= 82: return "Cat 1"
    if v <= 95: return "Cat 2"
    if v <= 112: return "Cat 3"
    if v <= 136: return "Cat 4"
    return "Cat 5"


async def fetch_hurricanes_v2():
    try:
        data = await _aget_json(NHC_URL,
                                headers={"User-Agent": NWS_UA, "Accept": "application/json"},
                                timeout=30, tries=2)
    except Exception:
        return _payload("hurricanes", [])
    items = []
    storms = data.get("activeStorms") or data.get("storms") or []
    for s in storms:
        lat = s.get("latitudeNumeric"); lng = s.get("longitudeNumeric")
        if lat is None or lng is None:
            continue
        items.append({
            "id": s.get("id") or s.get("binNumber") or s.get("name"),
            "lat": lat, "lng": lng,
            "label": s.get("name"),
            "classification": s.get("classification"),
            "category": _saffir_simpson(s.get("intensity")),
            "intensity_kts": s.get("intensity"),
            "pressure": s.get("pressure"),
            "movement_dir": s.get("movementDir"),
            "movement_speed": s.get("movementSpeed"),
            "last_update": s.get("lastUpdate"),
            "color": "#7B1FA2",
        })
    return _payload("hurricanes", items)


# ============================================================================
# GDACS multi-hazard
# ============================================================================
GDACS_URL = "https://www.gdacs.org/gdacsapi/api/events/geteventlist/SEARCH?"

_GDACS_TYPE_COLOR = {"EQ": "#FF9500", "TC": "#7B1FA2", "FL": "#1976D2",
                     "VO": "#FF6F00", "WF": "#FF6B00", "DR": "#8B7355"}
_GDACS_ALERT_COLOR = {"Green": "#76FF03", "Orange": "#FF9500", "Red": "#FF3D3D"}


async def fetch_gdacs():
    try:
        data = await _aget_json(GDACS_URL,
                                headers={"User-Agent": "globe-recon/1.0"},
                                timeout=20, tries=2)
    except Exception:
        return _payload("gdacs", [])
    items = []
    for feat in (data.get("features") or []):
        geom = feat.get("geometry") or {}
        props = feat.get("properties") or {}
        if geom.get("type") != "Point":
            continue
        coords = geom.get("coordinates") or []
        if len(coords) < 2:
            continue
        lng, lat = coords[0], coords[1]
        evt = props.get("eventtype") or "?"
        alert = props.get("alertlevel") or "Green"
        color = _GDACS_ALERT_COLOR.get(alert, _GDACS_TYPE_COLOR.get(evt, "#FFB300"))
        items.append({
            "id": str(props.get("eventid") or f"gdacs-{len(items)}"),
            "lat": lat, "lng": lng,
            "label": props.get("name") or evt,
            "event_type": evt,
            "alert_level": alert,
            "severity_score": props.get("severityvalue"),
            "country": props.get("country"),
            "from_date": props.get("fromdate"),
            "to_date": props.get("todate"),
            "url": (props.get("url") or {}).get("report") if isinstance(props.get("url"), dict) else None,
            "category": evt,
            "color": color,
        })
    return _payload("gdacs", items)


# ============================================================================
# Submarine cables (TeleGeography landing points)
# ============================================================================
SUBCABLE_URL = "https://www.submarinecablemap.com/api/v3/landing-point/landing-point-geo.json"


async def fetch_submarine_cables():
    try:
        data = await _aget_json(SUBCABLE_URL,
                                headers={"User-Agent": "globe-recon/1.0"},
                                timeout=20, tries=2)
    except Exception:
        return _payload("submarine_cables", [])
    items = []
    for feat in (data.get("features") or []):
        geom = feat.get("geometry") or {}
        if geom.get("type") != "Point":
            continue
        coords = geom.get("coordinates") or []
        if len(coords) < 2:
            continue
        lng, lat = coords[0], coords[1]
        props = feat.get("properties") or {}
        items.append({
            "id": str(props.get("id") or f"sc-{len(items)}"),
            "lat": lat, "lng": lng,
            "label": props.get("name") or "Submarine cable landing",
            "country": props.get("country"),
            "color": "#00BCD4",
        })
    return _payload("submarine_cables", items)


# ============================================================================
# Volcanoes — OSM Overpass natural=volcano
# ----------------------------------------------------------------------------
# 2026-05-24: Smithsonian GVP WFS started returning a Java ServiceException
# on every request (their server bug, not ours). Swapped to OSM Overpass
# which is the same source the collector already uses for nuclear /
# military_bases / power_plants and reliably returns 200 from both proxy
# providers. ~10k volcano nodes globally; ~5k are named. We tag each with
# the volcano type and last eruption when OSM has those keys (about half do).
# ============================================================================
VOLCANO_OVERPASS_QL = (
    '[out:json][timeout:90];'
    'node["natural"="volcano"];'
    'out;'
)


async def fetch_volcanoes():
    import urllib.parse
    url = ("https://overpass-api.de/api/interpreter?data="
           + urllib.parse.quote(VOLCANO_OVERPASS_QL))
    try:
        data = await _aget_json(
            url,
            headers={"User-Agent": "globe-recon/1.0",
                     "Accept": "application/json"},
            timeout=120, tries=2)
    except Exception:
        return _payload("volcanoes", [])
    items = []
    for el in data.get("elements", []):
        lat, lng = el.get("lat"), el.get("lon")
        if lat is None or lng is None:
            continue
        tags = el.get("tags") or {}
        name = tags.get("name") or tags.get("name:en")
        items.append({
            "id": f"osm-volcano-{el.get('id')}",
            "lat": lat, "lng": lng,
            "label": name or "Volcano",
            "country": tags.get("addr:country") or tags.get("is_in:country"),
            "primary_type": tags.get("volcano:type"),
            "last_known_eruption": tags.get("volcano:last_eruption"),
            "status": tags.get("volcano:status"),
            "elevation_m": tags.get("ele"),
            "wikipedia": tags.get("wikipedia"),
            "category": "Volcano",
            "color": "#FF6F00",
        })
    return _payload("volcanoes", items)


# ============================================================================
# OSM Overpass bbox helpers (used by power_plants + hospitals)
# ============================================================================
OVERPASS_ENDPOINTS = [
    "https://overpass-api.de/api/interpreter",
    "https://overpass.kumi.systems/api/interpreter",
]
OVERPASS_HDR = {
    "User-Agent": "globe-recon/1.0 (+osiris-globe-collector)",
    "Accept": "application/json",
    "Accept-Encoding": "gzip, deflate",
}

OSM_WORLD_BOXES = [
    (49, -140, 72, -52), (15, -125, 49, -85), (37, -85, 49, -60),
    (25, -85, 37, -60), (-5, -90, 25, -60), (-56, -82, -5, -34),
    (50, -10, 72, 35), (42, -10, 55, 20), (35, -10, 42, 30),
    (42, 20, 55, 40), (50, 28, 72, 68),
    (15, -18, 38, 42), (-5, -18, 15, 42), (-35, 25, -5, 52),
    (12, 32, 42, 62),
    (40, 55, 78, 135), (8, 68, 25, 88), (25, 68, 37, 88),
    (10, 88, 32, 108), (10, 108, 32, 120),
    (32, 100, 43, 115), (32, 115, 54, 135),
    (30, 129, 46, 146), (-10, 95, 25, 140), (20, 60, 40, 75),
    (-47, 112, -10, 180), (-25, -180, 25, -125),
]

# Hospital-density tuned bbox set. Tested with NYC bbox (0.5°×1°) which
# returned 157 hospitals in 6.3s — so 1°×1° boxes are the sweet spot for dense
# regions, and 5-10° for sparse regions. Total ~80 boxes — gather sequential
# (concurrency=1) with retries to dodge the DNS-burst issue + 504 storms.
# Sparse oceans get a single box; populated regions get 1°-2° boxes.
def _gen_hospital_boxes():
    boxes = []
    # 1) Dense regions: 1° × 1° lat × lng grids
    # US east coast metro corridor
    for lat in range(25, 49):
        for lng_start in range(-85, -60, 2):
            boxes.append((lat, lng_start, lat+1, lng_start+2))
    # Western Europe core
    for lat in range(42, 56):
        for lng_start in range(-10, 20, 2):
            boxes.append((lat, lng_start, lat+1, lng_start+2))
    # India
    for lat in range(8, 37, 2):
        for lng_start in range(68, 88, 4):
            boxes.append((lat, lng_start, lat+2, lng_start+4))
    # China + Korea
    for lat in range(20, 46, 3):
        for lng_start in range(100, 135, 5):
            boxes.append((lat, lng_start, lat+3, lng_start+5))
    # Japan
    boxes += [(30, 129, 36, 138), (36, 129, 42, 142), (42, 140, 46, 146)]
    # SE Asia archipelago
    for lat in range(-10, 22, 4):
        for lng_start in range(95, 145, 5):
            boxes.append((lat, lng_start, lat+4, lng_start+5))
    # 2) Medium regions: 5° × 5°
    # Canada + Alaska
    for lat in range(49, 70, 5):
        for lng_start in range(-140, -52, 12):
            boxes.append((lat, lng_start, lat+5, lng_start+12))
    # Central + South America
    for lat in range(-56, 25, 5):
        for lng_start in range(-90, -34, 12):
            boxes.append((lat, lng_start, lat+5, lng_start+12))
    # US central + west
    for lat in range(25, 49, 4):
        for lng_start in range(-125, -85, 8):
            boxes.append((lat, lng_start, lat+4, lng_start+8))
    # Northern Europe + Russia
    for lat in range(50, 72, 5):
        for lng_start in range(-10, 70, 15):
            boxes.append((lat, lng_start, lat+5, lng_start+15))
    # Africa
    for lat in range(-35, 38, 10):
        for lng_start in range(-20, 52, 15):
            boxes.append((lat, lng_start, lat+10, lng_start+15))
    # Middle East
    for lat in range(12, 42, 6):
        for lng_start in range(32, 65, 8):
            boxes.append((lat, lng_start, lat+6, lng_start+8))
    # Russia east + Central Asia
    boxes += [(40, 55, 60, 100), (40, 100, 60, 135), (60, 55, 78, 135)]
    # Pakistan + Afghanistan
    boxes += [(20, 60, 40, 75)]
    # Australia + NZ
    for lat in range(-47, -8, 8):
        for lng_start in range(112, 180, 12):
            boxes.append((lat, lng_start, lat+8, lng_start+12))
    # 3) Sparse: Pacific Islands
    boxes += [(-25, -180, 25, -125), (-25, 145, 0, 180)]
    return boxes


OSM_HOSPITAL_BOXES = _gen_hospital_boxes()


async def _overpass_one_box(union_clause: str, bbox: tuple):
    """Run one Overpass bbox query with retry-on-504 and exponential backoff.
    Overpass servers throttle hard under load; 504s clear after a few seconds."""
    s, w, n, e = bbox
    body = union_clause.format(s=s, w=w, n=n, e=e)
    ql = f'[out:json][timeout:30];({body});out center tags;'
    url = OVERPASS_ENDPOINTS[0] + "?data=" + urllib.parse.quote(ql, safe="")
    for attempt in range(4):
        try:
            r = await _F.aget(url, headers=OVERPASS_HDR, timeout=35)
            if r.status == 200 and r.body:
                try:
                    return json.loads(r.body).get("elements", []) or []
                except Exception:
                    return []
            # 504 / 429 = upstream overloaded — back off + retry
            if r.status in (429, 502, 503, 504):
                await asyncio.sleep(2 ** attempt + 1)
                continue
            return []
        except Exception:
            await asyncio.sleep(2 ** attempt)
    return []


async def _overpass_world(union_clause: str, boxes: list, concurrency: int = 2):
    sem = asyncio.Semaphore(concurrency)

    async def _one(b):
        async with sem:
            return await _overpass_one_box(union_clause, b)

    results = await asyncio.gather(*(_one(b) for b in boxes))
    seen = set(); out = []
    for r in results:
        for el in r:
            key = (el.get("type"), el.get("id"))
            if key in seen:
                continue
            seen.add(key); out.append(el)
    return {"elements": out}


_POWER_SRC_LABEL = {
    "solar": "Solar", "photovoltaic": "Solar", "wind": "Wind",
    "hydro": "Hydro", "water": "Hydro", "tidal": "Hydro",
    "gas": "Natural Gas", "natural_gas": "Natural Gas",
    "coal": "Coal", "lignite": "Coal", "oil": "Oil", "diesel": "Oil",
    "nuclear": "Nuclear", "biomass": "Biomass", "biogas": "Biomass",
    "geothermal": "Geothermal", "battery": "Battery",
    "energy_storage": "Battery", "waste": "Waste",
}

# Canonical fuel buckets used in the merged output. Sources (OSM tags, WRI
# `primary_fuel`, EIA Energy Source 1) feed in via _fuel_canonical().
_POWER_FUEL_COLOR = {
    "Nuclear":      "#7C4DFF",  # purple
    "Coal":         "#424242",  # dark grey
    "Natural Gas":  "#FF9800",  # orange
    "Oil":          "#5D4037",  # brown
    "Wind":         "#03A9F4",  # light blue
    "Solar":        "#FDD835",  # yellow
    "Hydro":        "#1E88E5",  # blue
    "Biomass":      "#8BC34A",  # green
    "Geothermal":   "#E64A19",  # red-orange
    "Battery":      "#9E9E9E",  # grey
    "Waste":        "#795548",  # brown
    "Other":        "#F5A623",  # amber (legacy power_plant color)
}

# Map raw fuel strings from various sources to canonical buckets.
_FUEL_NORMALIZE = {
    # OSM plant:source already maps via _POWER_SRC_LABEL
    # WRI primary_fuel values:
    "nuclear": "Nuclear",
    "coal": "Coal", "petcoke": "Coal", "lignite": "Coal",
    "gas": "Natural Gas", "natural_gas": "Natural Gas", "cogeneration": "Natural Gas",
    "oil": "Oil", "petroleum": "Oil", "diesel": "Oil",
    "wind": "Wind",
    "solar": "Solar", "photovoltaic": "Solar",
    "hydro": "Hydro", "water": "Hydro", "tidal": "Hydro", "wave_and_tidal": "Hydro",
    "biomass": "Biomass", "biogas": "Biomass",
    "geothermal": "Geothermal",
    "storage": "Battery", "battery": "Battery", "energy_storage": "Battery",
    "waste": "Waste",
    # EIA Energy Source 1 codes (subset; see EIA-923/860 instructions):
    "nuc": "Nuclear",
    "bit": "Coal", "sub": "Coal", "lig": "Coal", "ant": "Coal", "rc": "Coal",
    "wc": "Coal", "sgc": "Coal",
    "ng": "Natural Gas", "lfg": "Natural Gas", "obg": "Natural Gas",
    "pg": "Natural Gas", "bfg": "Natural Gas", "og": "Natural Gas",
    "dfo": "Oil", "rfo": "Oil", "wo": "Oil", "kero": "Oil", "jf": "Oil",
    "pc": "Coal", "sgp": "Oil",
    "wnd": "Wind",
    "sun": "Solar",
    "wat": "Hydro", "mwh": "Hydro",  # MWh = pumped storage hydro
    "wds": "Biomass", "wdl": "Biomass", "ab": "Biomass", "msw": "Biomass",
    "obs": "Biomass", "obl": "Biomass", "slw": "Biomass",
    "geo": "Geothermal",
    "msb": "Waste", "msn": "Waste", "tdf": "Waste",
    "blq": "Biomass",       # Black liquor (paper-mill cogen)
    "ker": "Oil",           # Kerosene
    "wh": "Other",          # Waste heat (industrial recovery)
    "pur": "Other",         # Purchased steam
    "oth": "Other",
}


def _fuel_canonical(raw: str | None) -> str:
    if not raw:
        return "Other"
    s = str(raw).strip().lower().split(";")[0].split(",")[0].strip()
    if not s:
        return "Other"
    # Try OSM-style label first (already canonical capitalization)
    if s in _POWER_SRC_LABEL:
        return _POWER_SRC_LABEL[s]
    if s in _FUEL_NORMALIZE:
        return _FUEL_NORMALIZE[s]
    # Title-cased pass for unknowns that already look like our buckets
    title = s.title()
    if title in _POWER_FUEL_COLOR:
        return title
    return "Other"


def _color_by_fuel(fuel: str) -> str:
    return _POWER_FUEL_COLOR.get(fuel, _POWER_FUEL_COLOR["Other"])


def _pp_dedup_key(lat: float, lng: float, operator: str | None) -> tuple:
    """Key for collapsing duplicate plants across sources.

    ~110m precision via 3-decimal rounding + first-20-char lowercased operator
    name. Operator is often missing/inconsistent, so when blank we fall back
    to a geo-only key — that does collapse co-located plants from the same
    multi-unit campus into one row, which is desired for this layer.
    """
    try:
        lat_r = round(float(lat), 3)
        lng_r = round(float(lng), 3)
    except (TypeError, ValueError):
        return ("__bad__", lat, lng)
    op = (operator or "").strip().lower()[:20]
    return (lat_r, lng_r, op)


POWER_PLANTS_CLAUSE = (
    'node["power"="plant"]({s},{w},{n},{e});'
    'way["power"="plant"]({s},{w},{n},{e});'
    'relation["power"="plant"]({s},{w},{n},{e});'
)


def _output_to_mw(s: str | None) -> float | None:
    """Parse OSM plant:output:electricity values into MW.

    Common forms: '12 MW', '1.5 GW', '500 kW', 'yes', or a bare number (assumed MW).
    """
    if not s:
        return None
    txt = str(s).strip().lower().replace(",", "")
    if txt in ("yes", "no", ""):
        return None
    mult = 1.0  # MW default
    for unit, m in (("gw", 1000.0), ("mw", 1.0), ("kw", 0.001), ("w", 1e-6)):
        if txt.endswith(unit):
            txt = txt[: -len(unit)].strip()
            mult = m
            break
    try:
        v = float(txt.split()[0]) if txt else None
    except (ValueError, IndexError):
        return None
    if v is None:
        return None
    return round(v * mult, 3)


def _normalize_power(raw):
    items = []
    for el in raw.get("elements", []):
        tags = el.get("tags") or {}
        lat, lng = el.get("lat"), el.get("lon")
        center = el.get("center")
        if (lat is None or lng is None) and isinstance(center, dict):
            lat, lng = center.get("lat"), center.get("lon")
        if lat is None or lng is None:
            continue
        src = (tags.get("plant:source") or "").lower().split(";")[0]
        fuel = _fuel_canonical(src)
        capacity_mw = _output_to_mw(tags.get("plant:output:electricity"))
        start_date = tags.get("start_date") or tags.get("opening_date")
        commissioning_year = None
        if start_date:
            try:
                commissioning_year = int(str(start_date)[:4])
            except ValueError:
                commissioning_year = None
        items.append({
            "id": f"osm-{el.get('type')}-{el.get('id')}",
            "lat": lat, "lng": lng,
            "label": tags.get("name") or _POWER_SRC_LABEL.get(src, "Power Plant"),
            "category": fuel,
            "primary_fuel": fuel,
            "source_type": src or None,
            "output": tags.get("plant:output:electricity"),
            "capacity_mw": capacity_mw,
            "operator": tags.get("operator"),
            "country": tags.get("addr:country") or tags.get("country"),
            "commissioning_year": commissioning_year,
            "color": _color_by_fuel(fuel),
            "sources": ["osm"],
        })
    return items


# ============================================================================
# Power plants — WRI Global Power Plant Database v1.3 (CC BY 4.0)
# ----------------------------------------------------------------------------
# ~30k plants worldwide, hand-curated. Provides capacity_mw, primary_fuel,
# owner, commissioning_year — the fields OSM consistently lacks. Fetched
# from the wri-dataportal-prod S3 bucket (CSV inside a zip). The dataset has
# not been refreshed since 2021 but the inventory is still the best public
# global baseline.
# ============================================================================
WRI_PP_URL = (
    "https://wri-dataportal-prod.s3.amazonaws.com/manual/"
    "global_power_plant_database_v_1_3.zip"
)


async def _pp_wri():
    """Fetch + parse WRI Global Power Plant Database. Returns normalized items."""
    import zipfile

    try:
        r = await _F.aget(WRI_PP_URL, headers={"Accept": "application/zip"},
                          timeout=90)
        if r.status != 200 or not r.body:
            print(f"power_plants: wri http={r.status} bytes={len(r.body or b'')}")
            return []
        zf = zipfile.ZipFile(io.BytesIO(r.body))
    except Exception as e:
        print(f"power_plants: wri fetch FAIL {type(e).__name__}: {e}")
        return []
    # Find the CSV inside the zip (single top-level CSV expected).
    csv_name = None
    for n in zf.namelist():
        if n.lower().endswith(".csv") and "global_power_plant_database" in n.lower():
            csv_name = n
            break
    if not csv_name:
        print(f"power_plants: wri zip missing csv (members={zf.namelist()[:5]})")
        return []
    try:
        with zf.open(csv_name) as fh:
            text = io.TextIOWrapper(fh, encoding="utf-8", errors="replace")
            reader = csv.DictReader(text)
            items = []
            for row in reader:
                try:
                    lat = float(row.get("latitude") or "")
                    lng = float(row.get("longitude") or "")
                except (TypeError, ValueError):
                    continue
                try:
                    cap = float(row["capacity_mw"]) if row.get("capacity_mw") else None
                except (TypeError, ValueError):
                    cap = None
                try:
                    year = int(float(row["commissioning_year"])) if row.get("commissioning_year") else None
                except (TypeError, ValueError):
                    year = None
                fuel = _fuel_canonical(row.get("primary_fuel"))
                items.append({
                    "id": f"wri-{(row.get('gppd_idnr') or '').strip() or len(items)}",
                    "lat": lat, "lng": lng,
                    "label": (row.get("name") or "").strip() or "Power Plant",
                    "category": fuel,
                    "primary_fuel": fuel,
                    "capacity_mw": cap,
                    "operator": (row.get("owner") or "").strip() or None,
                    "country": (row.get("country_long") or row.get("country") or "").strip() or None,
                    "commissioning_year": year,
                    "color": _color_by_fuel(fuel),
                    "sources": ["wri"],
                })
    except Exception as e:
        print(f"power_plants: wri parse FAIL {type(e).__name__}: {e}")
        return []
    print(f"power_plants: wri -> {len(items)} plants")
    return items


# ============================================================================
# Power plants — EIA Form 860 (US, ~12k plants, public domain)
# ----------------------------------------------------------------------------
# Two sheets joined per plant_code:
#   2___Plant_*.xlsx     -> Plant Code, Plant Name, State, Latitude, Longitude,
#                           Utility Name, Sector Name
#   3_1_Generator_*.xlsx -> per-generator Nameplate Capacity (MW), Technology,
#                           Energy Source 1, Operating Year, Planned Retirement
# We sum nameplate capacity per plant and pick the dominant fuel (max MW share).
#
# NOTE: www.eia.gov rejects ProxyRack residential CONNECT with HTTP 565
# (Akamai/anti-bot edge configured to reject residential ASNs). Production
# collector runs on GitHub Actions runners that have direct egress to .gov
# sites, so we attempt proxy first, then fall back to direct. This is a
# defensible carve-out: EIA is US government open data and the home IP is
# never exposed because this code paths runs in GH Actions.
# ============================================================================
EIA860_URLS = [
    # Current (2024 final, released Sep 2025) lives in /xls/. Older years move
    # to /archive/xls/. Walk from newest to oldest and use the first one that
    # downloads — keeps us self-healing across yearly cadence.
    "https://www.eia.gov/electricity/data/eia860/xls/eia8602024.zip",
    "https://www.eia.gov/electricity/data/eia860/xls/eia8602023ER.zip",
    "https://www.eia.gov/electricity/data/eia860/archive/xls/eia8602023.zip",
    "https://www.eia.gov/electricity/data/eia860/archive/xls/eia8602022.zip",
]


def _eia_fetch_bytes(url: str, timeout: int = 90) -> bytes | None:
    """Best-effort download of the EIA-860 zip.

    Try ProxyRack first; on 565/Akamai-reject, fall back to direct urllib.
    Direct is acceptable here because (a) EIA is US gov open data with no
    bot-block beyond their residential-ASN edge filter, and (b) production
    runs on GH Actions egress, not the user's home IP.
    """
    proxy = os.environ.get("PROXYRACK_PROXY_URL") or os.environ.get("HTTPS_PROXY")
    if proxy:
        try:
            handler = urllib.request.ProxyHandler({"http": proxy, "https": proxy})
            opener = urllib.request.build_opener(handler)
            opener.addheaders = [("User-Agent",
                                  "globe-recon/1.0 (+osiris-globe-collector)")]
            with opener.open(url, timeout=timeout) as r:
                if r.status == 200:
                    return r.read()
        except Exception as e:
            print(f"power_plants: eia proxy fetch failed ({type(e).__name__}: {e}); "
                  f"falling back to direct")
    # Direct fallback
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "globe-recon/1.0 (+osiris-globe-collector)"})
        with urllib.request.urlopen(req, timeout=timeout) as r:
            if r.status == 200:
                return r.read()
    except Exception as e:
        print(f"power_plants: eia direct fetch failed ({type(e).__name__}: {e}) "
              f"url={url}")
    return None


def _eia_pick_sheet(zf, prefix: str) -> tuple[str, bytes] | None:
    """Find the first xlsx whose filename starts with `prefix` (e.g. '2___Plant')."""
    for n in zf.namelist():
        base = n.rsplit("/", 1)[-1]
        if base.startswith(prefix) and base.lower().endswith(".xlsx"):
            return n, zf.read(n)
    return None


def _eia_load_xlsx_rows(xlsx_bytes: bytes, sheet_name: str | None = None):
    """Yield dict rows from an EIA xlsx. Headers are on the 2nd row (the 1st
    is a survey-year banner row); we skip until a row contains a known anchor
    column then use that as header."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    header = None
    for row in ws.iter_rows(values_only=True):
        if header is None:
            # EIA puts a one-cell banner row, then headers on row 2. Detect by
            # looking for a known column anchor.
            joined = "|".join(str(c) if c is not None else "" for c in row).lower()
            if "plant code" in joined or "utility id" in joined:
                header = [str(c).strip() if c is not None else "" for c in row]
            continue
        if all(c is None or c == "" for c in row):
            continue
        yield dict(zip(header, row))
    wb.close()


def _pp_eia860_parse(zip_bytes: bytes) -> list[dict]:
    import zipfile
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    plant_member = _eia_pick_sheet(zf, "2___Plant")
    gen_member = _eia_pick_sheet(zf, "3_1_Generator")
    if not plant_member:
        print(f"power_plants: eia zip missing 2___Plant sheet (members={zf.namelist()[:8]})")
        return []
    # Plants
    plants: dict[str, dict] = {}
    for row in _eia_load_xlsx_rows(plant_member[1]):
        code = row.get("Plant Code")
        if code is None:
            continue
        try:
            code_str = str(int(code))
        except (ValueError, TypeError):
            code_str = str(code).strip()
        try:
            lat = float(row.get("Latitude"))
            lng = float(row.get("Longitude"))
        except (TypeError, ValueError):
            continue
        plants[code_str] = {
            "id": f"eia-{code_str}",
            "lat": lat, "lng": lng,
            "label": (row.get("Plant Name") or "").strip() or f"EIA Plant {code_str}",
            "operator": (row.get("Utility Name") or "").strip() or None,
            "country": "United States",
            "state": (row.get("State") or "").strip() or None,
            "sector": (row.get("Sector Name") or "").strip() or None,
            "_fuel_mw": {},               # fuel -> summed MW (for dominant pick)
            "_capacity_mw": 0.0,
            "_min_year": None,
            "sources": ["eia860"],
        }
    # Generators — aggregate per plant
    if gen_member:
        for row in _eia_load_xlsx_rows(gen_member[1]):
            code = row.get("Plant Code")
            if code is None:
                continue
            try:
                code_str = str(int(code))
            except (ValueError, TypeError):
                code_str = str(code).strip()
            p = plants.get(code_str)
            if not p:
                continue
            try:
                mw = float(row.get("Nameplate Capacity (MW)") or 0)
            except (TypeError, ValueError):
                mw = 0.0
            fuel = _fuel_canonical(row.get("Energy Source 1"))
            if mw > 0:
                p["_capacity_mw"] += mw
                p["_fuel_mw"][fuel] = p["_fuel_mw"].get(fuel, 0.0) + mw
            yr = row.get("Operating Year")
            if yr:
                try:
                    yr_i = int(float(yr))
                    if p["_min_year"] is None or yr_i < p["_min_year"]:
                        p["_min_year"] = yr_i
                except (TypeError, ValueError):
                    pass
    # Finalize shape
    items = []
    for code_str, p in plants.items():
        # Dominant fuel = max-MW bucket; fallback Other if no generator data.
        fuel = "Other"
        if p["_fuel_mw"]:
            fuel = max(p["_fuel_mw"].items(), key=lambda kv: kv[1])[0]
        items.append({
            "id": p["id"],
            "lat": p["lat"], "lng": p["lng"],
            "label": p["label"],
            "category": fuel,
            "primary_fuel": fuel,
            "capacity_mw": round(p["_capacity_mw"], 3) if p["_capacity_mw"] else None,
            "operator": p["operator"],
            "country": p["country"],
            "state": p["state"],
            "sector": p["sector"],
            "commissioning_year": p["_min_year"],
            "color": _color_by_fuel(fuel),
            "sources": p["sources"],
        })
    return items


async def _pp_eia860():
    """Fetch + parse the most recent EIA-860 zip. Returns normalized items."""
    def _blocking() -> list[dict]:
        zip_bytes = None
        for url in EIA860_URLS:
            zip_bytes = _eia_fetch_bytes(url)
            if zip_bytes:
                print(f"power_plants: eia downloaded {len(zip_bytes)} bytes from {url}")
                break
        if not zip_bytes:
            print("power_plants: eia all sources failed")
            return []
        try:
            return _pp_eia860_parse(zip_bytes)
        except Exception as e:
            print(f"power_plants: eia parse FAIL {type(e).__name__}: {e}")
            return []

    items = await asyncio.to_thread(_blocking)
    print(f"power_plants: eia -> {len(items)} plants")
    return items


# ============================================================================
# Power plants — ENTSO-E (EU)
# ----------------------------------------------------------------------------
# Status: DOCUMENTED SKIP.
#
# We investigated the Transparency Platform Restful API for per-plant
# capacity. The available document types (A68 = Installed Generation
# Capacity per Unit, A71 = Generation Forecast, A75 = Actual Generation
# per Production Type) are *aggregated by production type per bidding
# zone* — not per-plant geographic records. The only per-unit doc (A95)
# requires the requesting party to register specific Production Unit
# EICs in advance and does not enumerate them.
#
# WRI v1.3 already provides ~6k EU power plants with capacity_mw,
# primary_fuel, owner and a non-trivial commissioning_year column. That
# is our EU per-plant baseline. ENTSO-E remains useful for the existing
# `eu_grid` layer (real-time outage signal, NOT inventory).
#
# The fetcher is kept as a no-op stub so the aggregator interface stays
# uniform; flipping to real data later means swapping this body without
# touching fetch_power_plants().
# ============================================================================
async def _pp_entsoe():
    """ENTSO-E does not expose per-plant capacity; skip with a one-line note."""
    print("power_plants: entsoe SKIP — Transparency Platform exposes only "
          "aggregated production-type capacity per bidding zone, not per-plant "
          "geo records. WRI covers EU inventory.")
    return []


# Cache last good per-source result so a transient failure in one source
# never zeroes the merged layer — same pattern as fetch_cctv's _CCTV_LAST_GOOD.
_PP_LAST_GOOD: dict[str, list] = {}


async def _pp_osm() -> list:
    """Existing OSM bbox-tiled fetcher, wrapped to match the per-source signature."""
    raw = await _overpass_world(POWER_PLANTS_CLAUSE, OSM_WORLD_BOXES, concurrency=4)
    return _normalize_power(raw)


def _pp_merge(sources: list[tuple[str, list]]) -> list:
    """Merge per-source items into a single deduped list.

    OSM is the geometric primary (its records keep ids unchanged). WRI/EIA
    records ENRICH a matching OSM record (filling capacity_mw, primary_fuel,
    operator, country, commissioning_year when OSM left them blank) and
    contribute to its `sources` list. When no OSM match exists, the alt
    source row is added verbatim.

    The merge order matters for collision precedence: OSM first, then WRI,
    then EIA. This means EIA wins the enrich for US plants (richer than WRI
    for US), WRI wins everywhere else.
    """
    by_key: dict = {}
    out: list = []
    # 1) Seed with OSM (primary records).
    osm_items = next((it for label, it in sources if label == "osm"), [])
    for it in osm_items:
        k = _pp_dedup_key(it["lat"], it["lng"], it.get("operator"))
        by_key[k] = it
        out.append(it)
    # 2) Layer in non-OSM sources in declared order.
    for label, items in sources:
        if label == "osm":
            continue
        for it in items:
            k = _pp_dedup_key(it["lat"], it["lng"], it.get("operator"))
            existing = by_key.get(k)
            if existing is None:
                # Geo-only fallback key (operator missing/different across sources
                # is the common case — e.g. WRI 'Pacific Gas & Electric Co' vs
                # OSM 'PG&E'). Try the lat/lng cell with empty operator before
                # accepting a new record.
                geo_k = (k[0], k[1], "")
                existing = by_key.get(geo_k)
            if existing is None:
                by_key[k] = it
                out.append(it)
                continue
            # Enrich existing — fill blank fields; never overwrite OSM ground truth.
            for field in ("capacity_mw", "primary_fuel", "operator", "country",
                          "commissioning_year", "label"):
                if not existing.get(field) and it.get(field):
                    existing[field] = it[field]
            # primary_fuel might be "Other" on OSM; prefer a non-Other value
            # from an enriching source.
            if existing.get("primary_fuel") == "Other" and it.get("primary_fuel") not in (None, "Other"):
                existing["primary_fuel"] = it["primary_fuel"]
                existing["category"] = it["primary_fuel"]
                existing["color"] = _color_by_fuel(it["primary_fuel"])
            srcs = existing.setdefault("sources", [])
            for s in it.get("sources", []):
                if s not in srcs:
                    srcs.append(s)
    # Final sweep: ensure color matches whatever fuel we ended with.
    for it in out:
        it["color"] = _color_by_fuel(it.get("primary_fuel") or "Other")
    return out


async def fetch_power_plants():
    """Merged power_plants layer: OSM (geometry primary) + WRI (global attrs)
    + EIA-860 (US per-plant). Each source isolated with its own try/except
    and a last-good cache so one source's failure never zeroes the layer."""

    async def _safe(label: str, coro):
        try:
            items = await coro
            n = len([r for r in items if r.get("lat") is not None
                     and r.get("lng") is not None])
            if n > 0:
                _PP_LAST_GOOD[label] = items
                print(f"power_plants: {label} -> {n} plants")
                return items
            cached = _PP_LAST_GOOD.get(label)
            if cached:
                print(f"power_plants: {label} -> 0 plants, reusing {len(cached)} cached")
                return cached
            print(f"power_plants: {label} -> 0 plants (no cache)")
            return []
        except Exception as e:
            cached = _PP_LAST_GOOD.get(label)
            if cached:
                print(f"power_plants: {label} FAIL {type(e).__name__}: {e}; "
                      f"reusing {len(cached)} cached")
                return cached
            print(f"power_plants: {label} FAIL {type(e).__name__}: {e}")
            return []

    osm = await _safe("osm", _pp_osm())
    wri = await _safe("wri", _pp_wri())
    eia = await _safe("eia860", _pp_eia860())
    entsoe = await _safe("entsoe", _pp_entsoe())  # currently no-op
    merged = _pp_merge([("osm", osm), ("wri", wri), ("eia860", eia),
                        ("entsoe", entsoe)])
    print(f"power_plants: merged -> {len(merged)} (osm={len(osm)} "
          f"wri={len(wri)} eia={len(eia)} entsoe={len(entsoe)})")
    return _payload("power_plants", merged)


HOSPITALS_CLAUSE = (
    'node["amenity"="hospital"]({s},{w},{n},{e});'
    'way["amenity"="hospital"]({s},{w},{n},{e});'
    'relation["amenity"="hospital"]({s},{w},{n},{e});'
)


def _safe_int(v):
    """Coerce OSM beds tag to int; OSM data is dirty ('120', '~50', '50+'),
    so strip non-digit suffixes and return None on failure."""
    if v is None:
        return None
    try:
        s = str(v).strip()
        if not s:
            return None
        digits = ""
        for ch in s:
            if ch.isdigit():
                digits += ch
            elif digits:
                break
        return int(digits) if digits else None
    except Exception:
        return None


def _normalize_hospitals(raw):
    items = []
    for i, el in enumerate(raw.get("elements", [])):
        tags = el.get("tags") or {}
        lat, lng = el.get("lat"), el.get("lon")
        center = el.get("center")
        if (lat is None or lng is None) and isinstance(center, dict):
            lat, lng = center.get("lat"), center.get("lon")
        if lat is None or lng is None:
            continue
        t = el.get("type")
        eid = el.get("id")
        items.append({
            "id": f"hosp-{t}-{eid}",
            "lat": lat, "lng": lng,
            "label": tags.get("name") or tags.get("name:en") or "Hospital",
            "operator": tags.get("operator"),
            "emergency": tags.get("emergency") == "yes",
            "beds": _safe_int(tags.get("beds")),
            "country": tags.get("addr:country") or tags.get("is_in:country"),
            "color": "#FF1744",
        })
    return items


async def _hospitals_one_box(bbox: tuple, endpoint: str):
    """Hospitals-tuned per-box Overpass runner. ProxyRack residential drops
    long-lived connections at ~16s, so Overpass server timeout is capped at
    25s and client at 28s. Retries on 504/429/connection-drop with short
    backoff. Returns [] on permanent failure (don't abort the world sweep)."""
    s, w, n, e = bbox
    body = HOSPITALS_CLAUSE.format(s=s, w=w, n=n, e=e)
    ql = f'[out:json][timeout:25];({body});out center tags;'
    url = endpoint + "?data=" + urllib.parse.quote(ql, safe="")
    for attempt in range(3):
        try:
            r = await _F.aget(url, headers=OVERPASS_HDR, timeout=28)
            if r.status == 200 and r.body:
                try:
                    return json.loads(r.body).get("elements", []) or []
                except Exception:
                    return []
            if r.status in (-1, 429, 502, 503, 504):
                await asyncio.sleep(1 + attempt)
                continue
            return []
        except Exception:
            await asyncio.sleep(1 + attempt)
    return []


async def fetch_hospitals():
    """Bbox-tiled Overpass — same shape as fetch_power_plants (per-box async
    + asyncio.Semaphore + (type,id) dedup) but uses the density-tuned
    OSM_HOSPITAL_BOXES split across both Overpass mirrors (main + kumi) to
    bypass overpass-api.de's 4-slot-per-IP limit. Round-robin assignment
    means concurrency=8 = 4 per server. Tighter per-box timeouts fit inside
    ProxyRack residential's ~16s connection ceiling on dense queries."""
    try:
        boxes = OSM_HOSPITAL_BOXES
        sem = asyncio.Semaphore(8)

        async def _one(i_b):
            i, b = i_b
            ep = OVERPASS_ENDPOINTS[i % len(OVERPASS_ENDPOINTS)]
            async with sem:
                return await _hospitals_one_box(b, ep)

        results = await asyncio.gather(*(_one(ib) for ib in enumerate(boxes)))
        seen = set()
        merged = []
        for r in results:
            for el in r:
                key = (el.get("type"), el.get("id"))
                if key in seen:
                    continue
                seen.add(key)
                merged.append(el)
        return _payload("hospitals", _normalize_hospitals({"elements": merged}))
    except Exception:
        return _payload("hospitals", [])


# ============================================================================
# Sanctioned vessels overlay — OpenSanctions IMOs ∩ live AIS
# ============================================================================
OPENSANCTIONS_CSV = "https://data.opensanctions.org/datasets/latest/maritime/maritime.csv"

_SANC_IMO_DETAIL: dict = {}
_SANC_CACHE_TIME = 0.0


async def _refresh_sanctioned_imos():
    global _SANC_IMO_DETAIL, _SANC_CACHE_TIME
    if _SANC_IMO_DETAIL and (time.time() - _SANC_CACHE_TIME) < 21600:
        return _SANC_IMO_DETAIL
    try:
        r = await _F.aget(OPENSANCTIONS_CSV,
                          headers={"Accept": "text/csv", "User-Agent": "globe-recon/1.0"},
                          timeout=30)
        if r.status != 200 or not r.body:
            return _SANC_IMO_DETAIL
        rows = csv.DictReader(io.StringIO(r.body.decode("utf-8", errors="ignore")))
        detail = {}
        for row in rows:
            imo = (row.get("imo") or "").strip()
            if not imo:
                continue
            if imo.upper().startswith("IMO"):
                imo = imo[3:]
            imo = imo.strip()
            if not imo or not imo.isdigit():
                continue
            detail[imo] = {
                "risk": row.get("risk") or "",
                "countries": row.get("countries") or "",
                "flag": row.get("flag") or "",
                "mmsi": row.get("mmsi") or "",
                "caption": row.get("caption") or "",
                "url": row.get("url") or "",
                "datasets": row.get("datasets") or "",
            }
        _SANC_IMO_DETAIL = detail
        _SANC_CACHE_TIME = time.time()
    except Exception:
        pass
    return _SANC_IMO_DETAIL


async def fetch_sanctioned_vessels():
    detail = await _refresh_sanctioned_imos()
    if not detail:
        return _payload("sanctioned_vessels", [])
    boats = await asyncio.to_thread(_read_latest_boats_blob)
    items = []
    for b in boats:
        imo = str(b.get("imo") or "").strip()
        if imo.upper().startswith("IMO"):
            imo = imo[3:]
        imo = imo.strip()
        if not imo:
            continue
        d = detail.get(imo)
        if not d:
            continue
        items.append({
            "id": f"sanc-{imo}",
            "lat": b.get("lat"), "lng": b.get("lng"),
            "label": b.get("name") or d.get("caption") or imo,
            "imo": imo,
            "mmsi": b.get("id") or d.get("mmsi"),
            "flag": d.get("flag") or b.get("flag"),
            "risk": d.get("risk"),
            "countries": d.get("countries"),
            "destination": b.get("destination"),
            "ship_type": b.get("ship_type"),
            "datasets": d.get("datasets"),
            "url": d.get("url"),
            "category": "Sanctioned",
            "color": "#FF1744",
        })
    return _payload("sanctioned_vessels", items)


# ============================================================================
# Port congestion — derived from stationary vessels in boats blob
# ============================================================================
async def fetch_port_congestion():
    boats = await asyncio.to_thread(_read_latest_boats_blob)
    if not boats:
        return _payload("port_congestion", [])
    grid = 0.5
    buckets: dict = {}
    for v in boats:
        try:
            spd = v.get("speed")
            spd_val = float(spd) if spd is not None else 0.0
        except (TypeError, ValueError):
            continue
        if spd_val > 1.0:
            continue
        lat, lng = v.get("lat"), v.get("lng")
        if not isinstance(lat, (int, float)) or not isinstance(lng, (int, float)):
            continue
        key = (round(lat / grid), round(lng / grid))
        buckets.setdefault(key, []).append(v)
    items = []
    for (gy, gx), vessels in buckets.items():
        if len(vessels) < 20:
            continue
        clat = sum(v["lat"] for v in vessels) / len(vessels)
        clng = sum(v["lng"] for v in vessels) / len(vessels)
        n = len(vessels)
        color = ("#7B1FA2" if n >= 100 else "#FF1744" if n >= 50
                 else "#FF9500" if n >= 30 else "#FFC400")
        items.append({
            "id": f"port-{gy}-{gx}",
            "lat": clat, "lng": clng,
            "label": f"{n} stationary vessels",
            "vessels": n,
            "category": "Port Congestion",
            "color": color,
        })
    items.sort(key=lambda i: -i["vessels"])
    return _payload("port_congestion", items[:300])


# ============================================================================
# NDBC ocean buoys (keyless text)
# ============================================================================
NDBC_URL = "https://www.ndbc.noaa.gov/data/latest_obs/latest_obs.txt"


def _ndbc_parse(text: str):
    items = []
    for line in text.splitlines():
        if not line or line.startswith("#"):
            continue
        parts = line.split()
        if len(parts) < 5:
            continue
        stn = parts[0]
        try:
            lat = float(parts[1]); lng = float(parts[2])
        except (ValueError, IndexError):
            continue
        if not (-90 <= lat <= 90 and -180 <= lng <= 180):
            continue

        def num(idx):
            try:
                v = parts[idx]
                if v in ("MM", "-"):
                    return None
                return float(v)
            except (IndexError, ValueError):
                return None
        wdir = num(8); wspd = num(9); gst = num(10)
        wvht = num(11); dpd = num(12)
        pres = num(15); atmp = num(17); wtmp = num(18)
        wh = wvht or 0
        color = ("#7B1FA2" if wh >= 6 else "#FF1744" if wh >= 4
                 else "#FF9500" if wh >= 2.5 else "#26C6DA"
                 if wh >= 1 else "#4FC3F7")
        items.append({
            "id": f"ndbc-{stn}",
            "lat": lat, "lng": lng,
            "label": stn,
            "station": stn,
            "wave_height_m": wvht,
            "wind_speed_ms": wspd,
            "gust_ms": gst,
            "wave_period_s": dpd,
            "pressure_hpa": pres,
            "air_temp_c": atmp,
            "sea_temp_c": wtmp,
            "wind_dir_deg": wdir,
            "category": "Ocean buoy",
            "color": color,
        })
    return items


async def fetch_ndbc_buoys():
    try:
        f = ProxyFetcher(impersonate="chrome146")
        r = await f.aget(NDBC_URL,
                         headers={"User-Agent": "globe-recon/1.0", "Accept": "text/plain"},
                         timeout=25)
        if r.status != 200 or not r.body:
            return _payload("ndbc_buoys", [])
        return _payload("ndbc_buoys", _ndbc_parse(r.body.decode("utf-8", errors="ignore")))
    except Exception:
        return _payload("ndbc_buoys", [])


# ============================================================================
# METAR airport weather (gzipped CSV global cache)
# ============================================================================
METAR_URL = "https://aviationweather.gov/data/cache/metars.cache.csv.gz"


def _metar_parse(csv_text: str):
    rdr = csv.reader(io.StringIO(csv_text))
    rows = list(rdr)
    header_idx = None
    for i, row in enumerate(rows):
        if row and row[0] == "raw_text":
            header_idx = i; break
    if header_idx is None:
        return []
    headers = rows[header_idx]
    col = {n: i for i, n in enumerate(headers)}

    def g(row, name):
        i = col.get(name); return row[i] if i is not None and i < len(row) else ""
    items = []
    for row in rows[header_idx + 1:]:
        if not row or len(row) < 5:
            continue
        try:
            lat = float(g(row, "latitude")); lng = float(g(row, "longitude"))
        except (ValueError, IndexError):
            continue
        if not (-90 <= lat <= 90 and -180 <= lng <= 180):
            continue
        sid = g(row, "station_id")

        def num(name):
            v = g(row, name)
            if not v:
                return None
            try: return float(v)
            except ValueError: return None
        fltcat = g(row, "flight_category") or "VFR"
        cat_color = {"VFR": "#76FF03", "MVFR": "#448AFF",
                     "IFR": "#FF1744", "LIFR": "#E040FB"}.get(fltcat, "#FFB300")
        items.append({
            "id": f"metar-{sid}",
            "lat": lat, "lng": lng,
            "label": f"{sid} · {fltcat}",
            "station": sid,
            "temp_c": num("temp_c"),
            "dewpoint_c": num("dewpoint_c"),
            "wind_dir_deg": num("wind_dir_degrees"),
            "wind_speed_kt": num("wind_speed_kt"),
            "gust_kt": num("wind_gust_kt"),
            "visibility": g(row, "visibility_statute_mi"),
            "flight_category": fltcat,
            "raw": g(row, "raw_text"),
            "track": num("wind_dir_degrees"),
            "category": fltcat,
            "color": cat_color,
        })
    return items


async def fetch_metar():
    try:
        f = ProxyFetcher(impersonate="chrome146")
        r = await f.aget(METAR_URL,
                         headers={"User-Agent": "globe-recon/1.0",
                                  "Accept-Encoding": "identity"},
                         timeout=25)
        if r.status != 200 or not r.body:
            return _payload("metar", [])
        try:
            csv_bytes = gzip.decompress(r.body)
        except Exception:
            csv_bytes = r.body
        return _payload("metar", _metar_parse(csv_bytes.decode("utf-8", errors="ignore")))
    except Exception:
        return _payload("metar", [])


# ============================================================================
# NOAA Tides — station catalog (real-time observation per station is N×slow,
# so we publish the catalog locations only; storm-surge polling deferred)
# ============================================================================
COOPS_STATIONS = "https://api.tidesandcurrents.noaa.gov/mdapi/prod/webapi/stations.json?type=waterlevels"

_TIDE_CATALOG_CACHE = None
_TIDE_CATALOG_TIME = 0.0


async def _tide_catalog():
    global _TIDE_CATALOG_CACHE, _TIDE_CATALOG_TIME
    if _TIDE_CATALOG_CACHE and (time.time() - _TIDE_CATALOG_TIME) < 86400:
        return _TIDE_CATALOG_CACHE
    try:
        data = await _aget_json(COOPS_STATIONS, timeout=20, tries=2)
        out = {}
        for s in (data.get("stations") or []):
            try:
                out[str(s.get("id"))] = (float(s.get("lat")), float(s.get("lng")),
                                          s.get("name") or s.get("id"))
            except (TypeError, ValueError):
                continue
        _TIDE_CATALOG_CACHE = out
        _TIDE_CATALOG_TIME = time.time()
        return out
    except Exception:
        return _TIDE_CATALOG_CACHE or {}


async def fetch_tides():
    cat = await _tide_catalog()
    if not cat:
        return _payload("tides", [])
    items = [{"id": f"tide-{sid}", "lat": lat, "lng": lng, "label": name,
              "station": sid, "category": "Tide Station", "color": "#0288D1"}
             for sid, (lat, lng, name) in cat.items()]
    return _payload("tides", items)


# ============================================================================
# Shipping lanes (static GitHub GeoJSON; LineStrings)
# ============================================================================
SHIPPING_LANES_URL = ("https://raw.githubusercontent.com/newzealandpaul/"
                      "Shipping-Lanes/main/data/Shipping_Lanes_v1.geojson")


async def fetch_shipping_lanes():
    """Returns shipping-lane segments. The source file uses MultiLineString
    geometry (3 features: Major / Middle / Minor lane groups). We explode each
    MultiLineString into individual LineString items keyed on the lane group,
    so the frontend can render them as polylines while still satisfying the
    point-shape contract (each item carries its first coord as lat/lng)."""
    try:
        data = await _aget_json(SHIPPING_LANES_URL,
                                headers={"User-Agent": "globe-recon/1.0"},
                                timeout=20, tries=2)
    except Exception:
        return _payload("shipping_lanes", [])
    items = []
    for feat in (data.get("features") or []):
        geom = feat.get("geometry") or {}
        gtype = geom.get("type")
        props = feat.get("properties") or {}
        lane_type = props.get("Type") or "Shipping"
        color = ("#FF9500" if lane_type == "Major"
                 else "#26A69A" if lane_type == "Middle"
                 else "#4FC3F7")
        # Each lane group is one MultiLineString containing many segments.
        if gtype == "MultiLineString":
            segments = geom.get("coordinates") or []
        elif gtype == "LineString":
            segments = [geom.get("coordinates") or []]
        else:
            continue
        for seg in segments:
            if not seg or len(seg[0]) < 2:
                continue
            lng, lat = seg[0][0], seg[0][1]
            items.append({
                "id": f"lane-{len(items)}",
                "lat": lat, "lng": lng,
                "label": f"{lane_type} Shipping Lane",
                "lane_type": lane_type,
                "geometry": {"type": "LineString", "coordinates": seg},
                "category": lane_type,
                "color": color,
            })
    return _payload("shipping_lanes", items)


# ============================================================================
# WHO Disease Outbreak News (replaces ProMED which is paywalled)
# ============================================================================
WHO_DON_URL = ("https://www.who.int/api/news/diseaseoutbreaknews"
               "?sf_format=json&$top=50&$orderby=PublicationDate%20desc")

# Country centroids (rough) so we can plot DON items globally without geocoding
_COUNTRY_CTR = {
    "AF": (33, 65), "AL": (41, 20), "DZ": (28, 3), "AO": (-12, 18),
    "AR": (-34, -64), "AU": (-25, 133), "AT": (47, 14), "BD": (24, 90),
    "BE": (51, 5), "BR": (-10, -55), "BG": (43, 25), "BF": (12, -2),
    "CM": (6, 13), "CA": (60, -95), "CF": (7, 21), "TD": (15, 19),
    "CL": (-30, -71), "CN": (35, 105), "CO": (4, -72), "CD": (0, 25),
    "CR": (10, -84), "CI": (8, -5), "HR": (45, 16), "CU": (22, -78),
    "CY": (35, 33), "CZ": (49, 15), "DK": (56, 10), "DJ": (12, 43),
    "DO": (19, -71), "EC": (-2, -78), "EG": (27, 30), "SV": (14, -89),
    "ER": (15, 39), "EE": (59, 26), "ET": (8, 38), "FI": (64, 26),
    "FR": (46, 2), "GA": (-1, 12), "GE": (42, 43), "DE": (51, 9),
    "GH": (8, -1), "GR": (39, 22), "GT": (16, -90), "GN": (10, -10),
    "HT": (19, -72), "HN": (15, -86), "HU": (47, 20), "IN": (20, 77),
    "ID": (-5, 120), "IR": (32, 53), "IQ": (33, 44), "IE": (53, -8),
    "IL": (31, 35), "IT": (43, 12), "JM": (18, -77), "JP": (36, 138),
    "JO": (31, 36), "KZ": (48, 68), "KE": (1, 38), "KP": (40, 127),
    "KR": (37, 128), "KW": (29, 47), "KG": (41, 75), "LA": (18, 105),
    "LV": (57, 25), "LB": (34, 36), "LR": (6, -9), "LY": (27, 17),
    "LT": (56, 24), "LU": (49, 6), "MG": (-20, 47), "MW": (-13, 34),
    "MY": (2, 112), "ML": (17, -4), "MT": (36, 14), "MR": (20, -12),
    "MU": (-20, 57), "MX": (23, -102), "MD": (47, 28), "MN": (46, 105),
    "MA": (32, -5), "MZ": (-18, 35), "MM": (22, 96), "NA": (-22, 17),
    "NP": (28, 84), "NL": (52, 5), "NZ": (-41, 174), "NI": (12, -85),
    "NE": (16, 8), "NG": (10, 8), "MK": (41, 21), "NO": (62, 10),
    "OM": (21, 57), "PK": (30, 70), "PA": (9, -80), "PG": (-6, 147),
    "PY": (-23, -58), "PE": (-10, -76), "PH": (13, 122), "PL": (52, 20),
    "PT": (39, -8), "QA": (25, 51), "RO": (46, 25), "RU": (60, 100),
    "RW": (-2, 30), "SA": (25, 45), "SN": (14, -14), "RS": (44, 21),
    "SL": (8, -11), "SG": (1, 103), "SK": (48, 19), "SI": (46, 14),
    "SO": (10, 49), "ZA": (-29, 24), "SS": (8, 30), "ES": (40, -4),
    "LK": (7, 81), "SD": (15, 30), "SR": (4, -56), "SE": (62, 15),
    "CH": (47, 8), "SY": (35, 38), "TW": (23, 121), "TJ": (39, 71),
    "TZ": (-6, 35), "TH": (15, 100), "TL": (-9, 126), "TG": (8, 1),
    "TT": (11, -61), "TN": (34, 9), "TR": (39, 35), "TM": (40, 60),
    "UG": (1, 32), "UA": (49, 32), "AE": (24, 54), "GB": (54, -2),
    "US": (40, -100), "UY": (-33, -56), "UZ": (41, 64), "VE": (8, -66),
    "VN": (16, 108), "YE": (15, 48), "ZM": (-15, 28), "ZW": (-19, 30),
}


# Country name → (lat, lng) for WHO DON geocoding. Built from the centroid
# table by mapping common WHO title spellings to ISO codes.
_COUNTRY_NAME_TO_CTR = {
    "afghanistan": "AF", "albania": "AL", "algeria": "DZ", "angola": "AO",
    "argentina": "AR", "australia": "AU", "austria": "AT", "bangladesh": "BD",
    "belgium": "BE", "brazil": "BR", "bulgaria": "BG", "burkina faso": "BF",
    "cameroon": "CM", "canada": "CA", "central african republic": "CF",
    "chad": "TD", "chile": "CL", "china": "CN", "colombia": "CO",
    "democratic republic of the congo": "CD", "congo": "CD",
    "costa rica": "CR", "côte d'ivoire": "CI", "ivory coast": "CI",
    "croatia": "HR", "cuba": "CU", "cyprus": "CY", "czech republic": "CZ",
    "denmark": "DK", "djibouti": "DJ", "dominican republic": "DO",
    "ecuador": "EC", "egypt": "EG", "el salvador": "SV", "eritrea": "ER",
    "estonia": "EE", "ethiopia": "ET", "finland": "FI", "france": "FR",
    "gabon": "GA", "georgia": "GE", "germany": "DE", "ghana": "GH",
    "greece": "GR", "guatemala": "GT", "guinea": "GN", "haiti": "HT",
    "honduras": "HN", "hungary": "HU", "india": "IN", "indonesia": "ID",
    "iran": "IR", "islamic republic of iran": "IR",
    "iraq": "IQ", "ireland": "IE", "israel": "IL", "italy": "IT",
    "jamaica": "JM", "japan": "JP", "jordan": "JO", "kazakhstan": "KZ",
    "kenya": "KE", "democratic people's republic of korea": "KP",
    "republic of korea": "KR", "south korea": "KR", "north korea": "KP",
    "kuwait": "KW", "kyrgyzstan": "KG", "lao people's democratic republic": "LA",
    "laos": "LA", "latvia": "LV", "lebanon": "LB", "liberia": "LR",
    "libya": "LY", "lithuania": "LT", "luxembourg": "LU", "madagascar": "MG",
    "malawi": "MW", "malaysia": "MY", "mali": "ML", "malta": "MT",
    "mauritania": "MR", "mauritius": "MU", "mexico": "MX", "moldova": "MD",
    "mongolia": "MN", "morocco": "MA", "mozambique": "MZ", "myanmar": "MM",
    "namibia": "NA", "nepal": "NP", "netherlands": "NL", "new zealand": "NZ",
    "nicaragua": "NI", "niger": "NE", "nigeria": "NG", "north macedonia": "MK",
    "norway": "NO", "oman": "OM", "pakistan": "PK", "panama": "PA",
    "papua new guinea": "PG", "paraguay": "PY", "peru": "PE", "philippines": "PH",
    "poland": "PL", "portugal": "PT", "qatar": "QA", "romania": "RO",
    "russia": "RU", "russian federation": "RU", "rwanda": "RW",
    "saudi arabia": "SA", "senegal": "SN", "serbia": "RS", "sierra leone": "SL",
    "singapore": "SG", "slovakia": "SK", "slovenia": "SI", "somalia": "SO",
    "south africa": "ZA", "south sudan": "SS", "spain": "ES", "sri lanka": "LK",
    "sudan": "SD", "suriname": "SR", "sweden": "SE", "switzerland": "CH",
    "syria": "SY", "syrian arab republic": "SY",
    "taiwan": "TW", "tajikistan": "TJ", "tanzania": "TZ", "thailand": "TH",
    "timor-leste": "TL", "togo": "TG", "trinidad and tobago": "TT",
    "tunisia": "TN", "turkey": "TR", "türkiye": "TR", "turkmenistan": "TM",
    "uganda": "UG", "ukraine": "UA", "united arab emirates": "AE",
    "uae": "AE", "united kingdom": "GB", "uk": "GB",
    "united states": "US", "united states of america": "US", "usa": "US",
    "uruguay": "UY", "uzbekistan": "UZ", "venezuela": "VE",
    "viet nam": "VN", "vietnam": "VN", "yemen": "YE", "zambia": "ZM",
    "zimbabwe": "ZW",
}


def _geocode_country(name: str):
    """Map a country name (free-form, lowercased, trimmed) to a (lat,lng)."""
    if not name:
        return None
    n = name.strip().lower().rstrip(".")
    if n in _COUNTRY_NAME_TO_CTR:
        return _COUNTRY_CTR[_COUNTRY_NAME_TO_CTR[n]]
    # Try the longest substring match (handles "DRC" → ...)
    for k, code in _COUNTRY_NAME_TO_CTR.items():
        if k in n or n in k:
            return _COUNTRY_CTR[code]
    return None


def _extract_countries_from_text(text: str):
    """Find ALL country names appearing in arbitrary text (title + summary).
    Sorted longest-name-first so 'Democratic Republic of the Congo' is matched
    before 'Congo'. Returns deduped list of (name, lat, lng) tuples."""
    if not text:
        return []
    t = text.lower()
    matches = []
    seen = set()
    # Sort keys by length DESC so multi-word countries win over substrings
    for name in sorted(_COUNTRY_NAME_TO_CTR.keys(), key=lambda s: -len(s)):
        if name in t:
            code = _COUNTRY_NAME_TO_CTR[name]
            if code in seen:
                continue
            seen.add(code)
            lat, lng = _COUNTRY_CTR[code]
            matches.append((name, lat, lng))
            # Remove from text so a substring of this longer name doesn't re-match
            t = t.replace(name, " ")
    return matches


async def fetch_who_don():
    """WHO DON outbreaks geocoded by every country mention found in title +
    summary. Multi-country outbreaks (e.g. cross-border Ebola) emit a marker
    per country. Longest-name-first matching handles 'DRC' vs 'Congo'."""
    try:
        url = WHO_DON_URL.replace("$top", "%24top").replace("$orderby", "%24orderby")
        data = await _aget_json(url, headers={"User-Agent": "globe-recon/1.0"},
                                timeout=20, tries=2)
    except Exception:
        return _payload("who_outbreaks", [])
    items = []
    rows = data.get("value") or data.get("d") or []
    for row in rows[:50]:
        title = row.get("Title") or row.get("title") or ""
        if not title:
            continue
        summary = row.get("Summary") or row.get("summary") or ""
        # Search BOTH title and summary for country mentions
        search_text = title + " " + summary[:500]
        countries = _extract_countries_from_text(search_text)
        if not countries:
            continue
        rid = row.get("Id") or row.get("id") or f"who-{len(items)}"
        pub = row.get("PublicationDate") or row.get("publicationDate")
        url_path = row.get("ItemDefaultUrl") or ""
        full_url = ("https://www.who.int" + url_path) if url_path.startswith("/") else url_path
        for cname, lat, lng in countries:
            items.append({
                "id": f"{rid}-{cname[:8]}",
                "lat": lat, "lng": lng,
                "label": title,
                "country_name": cname.title(),
                "publication_date": pub,
                "summary": summary[:300],
                "url": full_url,
                "category": "Disease Outbreak",
                "color": "#FF1744",
            })
    return _payload("who_outbreaks", items)


# ============================================================================
# iNaturalist research-grade observations (keyless)
# ============================================================================
async def fetch_inaturalist():
    """200 most-recent research-grade observations globally. iNat focuses on
    citizen-science wildlife sightings — complements GBIF (which can lag)."""
    url = ("https://api.inaturalist.org/v1/observations"
           "?per_page=200&order_by=created_at&order=desc"
           "&geo=true&quality_grade=research")
    try:
        data = await _aget_json(url, headers={"User-Agent": "globe-recon/1.0"},
                                timeout=30, tries=2)
    except Exception:
        return _payload("inaturalist", [])
    items = []
    for r in (data.get("results") or []):
        geo = r.get("geojson") or {}
        coords = geo.get("coordinates") or []
        if len(coords) < 2:
            continue
        lng, lat = coords[0], coords[1]
        taxon = r.get("taxon") or {}
        items.append({
            "id": f"inat-{r.get('id')}",
            "lat": lat, "lng": lng,
            "label": taxon.get("preferred_common_name") or taxon.get("name") or "Sighting",
            "scientific_name": taxon.get("name"),
            "rank": taxon.get("rank"),
            "iconic_taxon": taxon.get("iconic_taxon_name"),
            "observed_on": r.get("observed_on"),
            "url": r.get("uri"),
            "category": taxon.get("iconic_taxon_name") or "Wildlife",
            "color": "#26A69A",
        })
    return _payload("inaturalist", items)


# ============================================================================
# SEC EDGAR — recent 8-K (material event) filings, keyless
# ============================================================================
async def fetch_sec_edgar():
    """Recent 8-K material event filings. Plotted at SEC HQ since most filings
    don't expose the company's HQ address in the search-index response."""
    # Last 5 days, 100 most recent 8-K filings
    since = (datetime.now(timezone.utc) - timedelta(days=5)).strftime("%Y-%m-%d")
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    url = (f"https://efts.sec.gov/LATEST/search-index?q=&forms=8-K"
           f"&dateRange=custom&startdt={since}&enddt={today}")
    try:
        data = await _aget_json(url,
                                headers={"User-Agent": "globe-recon zac@zacstern.co",
                                         "Accept": "application/json"},
                                timeout=20, tries=2)
    except Exception:
        return _payload("sec_filings", [])
    hits = ((data.get("hits") or {}).get("hits")) or []
    items = []
    # SEC HQ centroid (Washington DC)
    sec_lat, sec_lng = 38.9, -77.04
    for h in hits[:100]:
        src = h.get("_source") or {}
        display_names = src.get("display_names") or []
        company = display_names[0] if display_names else "?"
        form = src.get("form") or "8-K"
        items.append({
            "id": h.get("_id") or f"sec-{len(items)}",
            "lat": sec_lat, "lng": sec_lng,
            "label": f"{company} · {form}",
            "company": company,
            "form_type": form,
            "filed_date": src.get("file_date"),
            "ciks": src.get("ciks"),
            "url": f"https://www.sec.gov/Archives/edgar/data/{src.get('ciks',['0'])[0]}/{(h.get('_id','').replace('-','').split(':')[0])}",
            "category": "Corporate Filing",
            "color": "#9C27B0",
        })
    return _payload("sec_filings", items)


# ============================================================================
# GBIF wildlife (keyless)
# ============================================================================
async def fetch_gbif():
    """Recent geotagged wildlife occurrences from GBIF. Pulls last 24h slice."""
    since = (datetime.now(timezone.utc) - timedelta(days=1)).strftime("%Y-%m-%d")
    until = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    url = (f"https://api.gbif.org/v1/occurrence/search"
           f"?hasCoordinate=true&hasGeospatialIssue=false"
           f"&lastInterpreted={since},{until}"
           f"&limit=300&offset=0")
    try:
        data = await _aget_json(url, headers={"User-Agent": "globe-recon/1.0"},
                                timeout=25, tries=2)
    except Exception:
        return _payload("wildlife_gbif", [])
    items = []
    for r in (data.get("results") or []):
        lat = r.get("decimalLatitude"); lng = r.get("decimalLongitude")
        if lat is None or lng is None:
            continue
        items.append({
            "id": str(r.get("gbifID") or r.get("occurrenceID") or len(items)),
            "lat": lat, "lng": lng,
            "label": r.get("scientificName") or r.get("species") or "Wildlife",
            "kingdom": r.get("kingdom"),
            "family": r.get("family"),
            "species": r.get("species"),
            "country": r.get("countryCode"),
            "event_date": r.get("eventDate"),
            "category": r.get("kingdom") or "Wildlife",
            "color": "#76FF03",
        })
    return _payload("wildlife_gbif", items)


# ============================================================================
# Env-gated layers — activate on key drop-in
# ============================================================================
ACLED_KEY = os.environ.get("ACLED_API_KEY", "")
ACLED_EMAIL = os.environ.get("ACLED_EMAIL", "")


async def fetch_acled():
    if not ACLED_KEY or not ACLED_EMAIL:
        return _payload("acled", [])
    since = (datetime.now(timezone.utc) - timedelta(days=14)).strftime("%Y-%m-%d")
    url = (f"https://api.acleddata.com/acled/read?key={ACLED_KEY}&email={ACLED_EMAIL}"
           f"&_format=json&limit=5000&event_date={since}")
    try:
        data = await _aget_json(url, timeout=30, tries=2)
    except Exception:
        return _payload("acled", [])
    items = []
    for e in (data.get("data") or []):
        try:
            lat = float(e.get("latitude")); lng = float(e.get("longitude"))
        except (TypeError, ValueError):
            continue
        etype = e.get("event_type") or "Event"
        items.append({
            "id": str(e.get("event_id_cnty") or len(items)),
            "lat": lat, "lng": lng,
            "label": e.get("notes") or etype,
            "event_type": etype,
            "sub_event_type": e.get("sub_event_type"),
            "actor1": e.get("actor1"),
            "actor2": e.get("actor2"),
            "date": e.get("event_date"),
            "country": e.get("country"),
            "fatalities": e.get("fatalities"),
            "category": etype,
            "color": ("#FF1744" if "battle" in etype.lower() or "violence" in etype.lower()
                      else "#FF6B00" if "explosion" in etype.lower() or "remote" in etype.lower()
                      else "#FFB300"),
        })
    return _payload("acled", items)


EIA_KEY = os.environ.get("EIA_API_KEY", "")
_EIA_BA = {
    "CISO": (37.0, -120.0, "California ISO"), "ERCO": (31.0, -99.0, "Texas (ERCOT)"),
    "PJM": (39.5, -77.5, "PJM Interconnection"), "MISO": (40.0, -90.0, "MISO"),
    "NYIS": (43.0, -75.5, "New York ISO"), "ISNE": (43.5, -71.5, "ISO New England"),
    "BPAT": (45.7, -120.5, "Bonneville Power Admin"), "FPL": (28.0, -81.5, "Florida Power & Light"),
    "SOCO": (33.0, -84.5, "Southern Company"), "DUK": (35.5, -80.5, "Duke Energy"),
    "TVA": (35.8, -86.5, "TVA"), "AECI": (37.5, -93.0, "AECI"),
    "PSCO": (39.5, -105.5, "Colorado Public Service"), "AZPS": (33.5, -112.0, "Arizona Public Service"),
}


# Pretty-print labels + colors for the fuel codes EIA emits.
_EIA_FUEL_META = {
    "COL": ("Coal", "#5C4033"),
    "NG":  ("Natural Gas", "#F5A623"),
    "NUC": ("Nuclear", "#76FF03"),
    "WAT": ("Hydro", "#26C6DA"),
    "WND": ("Wind", "#4FC3F7"),
    "SUN": ("Solar", "#FFEB3B"),
    "SNB": ("Solar (non-billing)", "#FFD600"),
    "GEO": ("Geothermal", "#FF6F00"),
    "OIL": ("Oil", "#3E2723"),
    "BAT": ("Battery", "#9C27B0"),
    "PS":  ("Pumped Storage", "#0288D1"),
    "OTH": ("Other", "#888"),
}


async def fetch_eia_grid():
    """Enrich each US Balancing Authority with: latest demand (MW) + latest
    generation by fuel type + clean-vs-fossil share. Two EIA endpoints called
    concurrently, joined by respondent code."""
    if not EIA_KEY:
        return _payload("eia_grid", [])
    base = "https://api.eia.gov/v2/electricity/rto"
    demand_url = (f"{base}/region-data/data/?api_key={EIA_KEY}"
                  f"&frequency=hourly&data[0]=value&facets[type][]=D"
                  f"&sort[0][column]=period&sort[0][direction]=desc&length=300")
    fuel_url = (f"{base}/fuel-type-data/data/?api_key={EIA_KEY}"
                f"&frequency=hourly&data[0]=value"
                f"&sort[0][column]=period&sort[0][direction]=desc&length=2000")
    try:
        demand_data, fuel_data = await asyncio.gather(
            _aget_json(demand_url, timeout=30, tries=2),
            _aget_json(fuel_url, timeout=30, tries=2),
        )
    except Exception:
        return _payload("eia_grid", [])
    demand_rows = ((demand_data.get("response") or {}).get("data")) or []
    fuel_rows = ((fuel_data.get("response") or {}).get("data")) or []
    demand_by_ba: dict = {}
    for r in demand_rows:
        ba = r.get("respondent")
        if ba and ba not in demand_by_ba and ba in _EIA_BA:
            demand_by_ba[ba] = r
    fuel_by_ba: dict = {}
    for r in fuel_rows:
        ba = r.get("respondent")
        ft = r.get("fueltype")
        if not ba or not ft or ba not in _EIA_BA:
            continue
        key = (ba, ft)
        if key in fuel_by_ba:
            continue
        try:
            mw = float(r.get("value")) if r.get("value") is not None else 0.0
        except (TypeError, ValueError):
            mw = 0.0
        if mw <= 0:
            continue
        fuel_by_ba[key] = mw
    items = []
    for ba, (lat, lng, label) in _EIA_BA.items():
        r = demand_by_ba.get(ba)
        mix = {ft: mw for (b, ft), mw in fuel_by_ba.items() if b == ba}
        total_gen = sum(mix.values()) or 0
        clean = sum(mix.get(ft, 0) for ft in ("NUC", "WAT", "WND", "SUN", "SNB", "GEO"))
        fossil = sum(mix.get(ft, 0) for ft in ("COL", "NG", "OIL"))
        clean_pct = (100 * clean / total_gen) if total_gen > 0 else None
        mix_breakdown = sorted(
            ((ft, mw, _EIA_FUEL_META.get(ft, (ft, "#888"))[0]) for ft, mw in mix.items()),
            key=lambda x: -x[1])
        dominant_ft = mix_breakdown[0][0] if mix_breakdown else None
        dominant_label = _EIA_FUEL_META.get(dominant_ft, (dominant_ft or "?", "#888"))[0]
        dominant_color = _EIA_FUEL_META.get(dominant_ft, (None, "#F5A623"))[1]
        items.append({
            "id": f"eia-{ba}",
            "lat": lat, "lng": lng,
            "label": (f"{label}: {int(float(r.get('value') or 0))} MW · "
                      f"{clean_pct:.0f}% clean" if r and clean_pct is not None
                      else label),
            "ba": ba,
            "demand_mw": r.get("value") if r else None,
            "period": r.get("period") if r else None,
            "total_generation_mw": int(total_gen) if total_gen else None,
            "clean_pct": round(clean_pct, 1) if clean_pct is not None else None,
            "clean_mw": int(clean) if clean else 0,
            "fossil_mw": int(fossil) if fossil else 0,
            "dominant_fuel": dominant_label,
            "fuel_mix": [
                {"fuel": ft, "label": lbl, "mw": int(mw)}
                for ft, mw, lbl in mix_breakdown
            ],
            "category": dominant_label,
            "color": dominant_color,
        })
    return _payload("eia_grid", items)


# ===========================================================================
# FRED macros — env-gated, free key at fredaccount.stlouisfed.org
# ===========================================================================
FRED_KEY = os.environ.get("FRED_API_KEY", "")
_FRED_HQ_LAT, _FRED_HQ_LNG = 38.6270, -90.1994

_FRED_SERIES = [
    ("UNRATE",     "US Unemployment Rate",      "#FF1744"),
    ("CPIAUCSL",   "US CPI (urban)",            "#FF9500"),
    ("DGS10",      "10-Year Treasury Yield",    "#1976D2"),
    ("DGS2",       "2-Year Treasury Yield",     "#0288D1"),
    ("T10Y2Y",     "10Y-2Y Spread",             "#7B1FA2"),
    ("DFF",        "Fed Funds Rate",            "#388E3C"),
    ("ICSA",       "Initial Jobless Claims",    "#FFEB3B"),
    ("PAYEMS",     "Nonfarm Payrolls",          "#26A69A"),
    ("UMCSENT",    "U-Mich Consumer Sentiment", "#9C27B0"),
    ("HOUST",      "Housing Starts",            "#5D4037"),
    ("VIXCLS",     "VIX",                       "#E91E63"),
    ("DCOILWTICO", "WTI Crude Oil",             "#3E2723"),
    ("DEXUSEU",    "USD/EUR",                   "#2196F3"),
    ("M2SL",       "M2 Money Stock",            "#FF6F00"),
]


async def fetch_fred_macros():
    """Top US macro indicators. Env-gated; activates on FRED_API_KEY."""
    if not FRED_KEY:
        return _payload("fred_macros", [])
    sem = asyncio.Semaphore(4)

    async def _one(series_id: str, label: str, color: str):
        async with sem:
            url = (f"https://api.stlouisfed.org/fred/series/observations"
                   f"?series_id={series_id}&api_key={FRED_KEY}&file_type=json"
                   f"&sort_order=desc&limit=2")
            try:
                d = await _aget_json(url, timeout=15, tries=2)
            except Exception:
                return None
            obs = d.get("observations") or []
            if not obs:
                return None
            latest = obs[0]
            prev = obs[1] if len(obs) > 1 else None
            try:
                val = float(latest.get("value")) if latest.get("value") not in (None, ".") else None
                prev_val = float(prev.get("value")) if prev and prev.get("value") not in (None, ".") else None
            except (TypeError, ValueError):
                return None
            if val is None:
                return None
            delta = (val - prev_val) if prev_val is not None else None
            return {
                "id": f"fred-{series_id}",
                "lat": _FRED_HQ_LAT, "lng": _FRED_HQ_LNG,
                "label": f"{label}: {val:.2f}",
                "series_id": series_id,
                "value": val,
                "previous_value": prev_val,
                "delta": delta,
                "delta_pct": (100 * delta / prev_val) if prev_val else None,
                "date": latest.get("date"),
                "previous_date": prev.get("date") if prev else None,
                "category": "Macro",
                "color": color,
            }

    results = await asyncio.gather(*(_one(sid, lbl, clr) for sid, lbl, clr in _FRED_SERIES))
    items = [r for r in results if r]
    # Spread markers radially around FRED HQ so they don't all stack.
    import math as _math
    for i, item in enumerate(items):
        angle = 2 * _math.pi * i / max(1, len(items))
        item["lat"] = _FRED_HQ_LAT + 0.5 * _math.sin(angle)
        item["lng"] = _FRED_HQ_LNG + 0.7 * _math.cos(angle)
    return _payload("fred_macros", items)


CLOUDFLARE_TOKEN = os.environ.get("CLOUDFLARE_RADAR_TOKEN", "")


async def fetch_internet_outages():
    if not CLOUDFLARE_TOKEN:
        return _payload("internet_outages", [])
    # CF Radar requires dateStart AND dateEnd in strict "YYYY-mm-ddTHH:MM:ssZ"
    # format — .isoformat() adds microseconds + offset which it rejects.
    since = (datetime.now(timezone.utc) - timedelta(days=30)).strftime("%Y-%m-%dT%H:%M:%SZ")
    until = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    url = (f"https://api.cloudflare.com/client/v4/radar/annotations/outages"
           f"?dateStart={since}&dateEnd={until}&format=json&limit=200")
    try:
        data = await _aget_json(url, headers={"Authorization": f"Bearer {CLOUDFLARE_TOKEN}"},
                                timeout=20, tries=2)
    except Exception:
        return _payload("internet_outages", [])
    items = []
    for r in (data.get("result", {}).get("annotations") or []):
        # CF returns `locations` as a list of ISO codes (one event can hit
        # multiple countries). Emit one marker per country.
        locs = r.get("locations") or []
        if not locs:
            locs = [(r.get("locationCode") or "")]
        for cc in locs:
            if not cc:
                continue
            ctr = _COUNTRY_CTR.get(cc)
            if not ctr:
                continue
            outage_info = r.get("outage") or {}
            items.append({
                "id": f"co-{r.get('id')}-{cc}",
                "lat": ctr[0], "lng": ctr[1],
                "label": r.get("description") or r.get("eventType") or "Internet outage",
                "country_code": cc,
                "event_type": r.get("eventType"),
                "outage_cause": outage_info.get("outageCause"),
                "outage_type": outage_info.get("outageType"),
                "started": r.get("startDate"),
                "ended": r.get("endDate"),
                "scope": r.get("scope"),
                "url": r.get("linkedUrl"),
                "category": outage_info.get("outageCause") or "Internet Outage",
                "color": "#9C27B0",
            })
    return _payload("internet_outages", items)


EBIRD_KEY = os.environ.get("EBIRD_API_KEY", "")


async def fetch_ebird():
    """Notable bird observations — env-gated. Pulls notable obs across ~30
    high-activity countries to stay under the 1000 req/day budget."""
    if not EBIRD_KEY:
        return _payload("ebird", [])
    regions = ["US", "CA", "MX", "BR", "AR", "CO", "EC", "PE", "GB", "DE",
               "FR", "ES", "IT", "PT", "NL", "SE", "NO", "FI", "RU", "JP",
               "CN", "IN", "AU", "NZ", "ZA", "KE", "TZ", "CR", "PA", "ID"]
    items = []
    sem = asyncio.Semaphore(4)

    async def _one(reg):
        async with sem:
            url = (f"https://api.ebird.org/v2/data/obs/{reg}/recent/notable"
                   f"?back=1&maxResults=50")
            try:
                data = await _aget_json(url,
                                        headers={"x-ebirdapitoken": EBIRD_KEY,
                                                 "User-Agent": "globe-recon/1.0"},
                                        timeout=15, tries=1)
                if isinstance(data, list):
                    return data
            except Exception:
                return []
            return []

    results = await asyncio.gather(*(_one(r) for r in regions))
    for region_results in results:
        for obs in region_results:
            lat = obs.get("lat"); lng = obs.get("lng")
            if lat is None or lng is None:
                continue
            items.append({
                "id": f"ebird-{obs.get('subId') or obs.get('speciesCode')}-{len(items)}",
                "lat": lat, "lng": lng,
                "label": obs.get("comName") or obs.get("sciName"),
                "species_code": obs.get("speciesCode"),
                "common_name": obs.get("comName"),
                "scientific_name": obs.get("sciName"),
                "location": obs.get("locName"),
                "how_many": obs.get("howMany"),
                "obs_date": obs.get("obsDt"),
                "category": "Bird sighting",
                "color": "#26C6DA",
            })
    return _payload("ebird", items)


GFW_TOKEN = os.environ.get("GFW_API_KEY", "") or os.environ.get("GFW_API_TOKEN", "")


# Curated high-interest maritime regions for SAR dark-vessel monitoring.
# GFW's 4wings/report endpoint REQUIRES a region-id (no global aggregate).
# Each entry: (label, lat, lng for marker, EEZ region-id). Region IDs from
# the public-eez-areas dataset — found via probing the dataset config.
# Coverage spans sanction-risk + illegal-fishing hotspots.
GFW_REGIONS = [
    ("Persian Gulf (Iran EEZ)",   26.0, 52.0, 8385),   # Iran
    ("North Korea EEZ",           39.5, 130.5, 8327),  # DPRK
    ("Russia Pacific EEZ",        55.0, 145.0, 8385),  # placeholder; will resolve
    ("South China Sea (China)",   18.0, 115.0, 8332),
    ("Venezuela EEZ",             10.0, -67.0, 8377),
    ("Libya EEZ",                 33.0, 19.0, 8307),
    ("Yemen EEZ",                 13.0, 48.0, 8331),
    ("Somali Coast",               2.0, 47.0, 8392),
]


async def fetch_gfw_sar():
    """GFW SAR dark-vessel detection AGGREGATE counts per region.

    GFW's 4wings/report endpoint serves grid-aggregated SAR detection counts
    NOT raw point detections (point-level data requires the heatmap MVT tile
    endpoint, which would need protobuf decoding). We query a curated set of
    high-interest maritime regions and emit one marker per region with the
    aggregate count for the last 24h. Off-cycle regions may return null —
    that's normal SAR pass cadence (Sentinel-1 revisits every 6-12 days).
    """
    if not GFW_TOKEN:
        return _payload("dark_vessels", [])
    yesterday = (datetime.now(timezone.utc) - timedelta(days=7)).strftime("%Y-%m-%d")
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    sem = asyncio.Semaphore(3)

    async def _one(region_id, lat, lng, label):
        async with sem:
            url = ("https://gateway.api.globalfishingwatch.org/v3/4wings/report"
                   f"?datasets[0]=public-global-sar-presence:latest"
                   f"&date-range={yesterday},{today}"
                   f"&spatial-resolution=LOW&temporal-resolution=DAILY"
                   f"&format=JSON&group-by=VESSEL_ID"
                   f"&region-id={region_id}&region-dataset=public-eez-areas")
            try:
                data = await _aget_json(
                    url, headers={"Authorization": f"Bearer {GFW_TOKEN}",
                                  "User-Agent": "globe-recon/1.0"},
                    timeout=20, tries=1)
            except Exception:
                return None
            # Response shape: {total, entries: [{dataset_id: count_or_null}]}
            entries = data.get("entries") or []
            cnt = 0
            for e in entries:
                for v in (e.values() if isinstance(e, dict) else []):
                    if v is not None and isinstance(v, (int, float)):
                        cnt += int(v)
            return (region_id, lat, lng, label, cnt, data.get("total"))

    results = await asyncio.gather(*(_one(r, la, ln, lb)
                                     for lb, la, ln, r in GFW_REGIONS))
    items = []
    for res in results:
        if not res:
            continue
        region_id, lat, lng, label, cnt, total = res
        items.append({
            "id": f"gfw-{region_id}",
            "lat": lat, "lng": lng,
            "label": f"{label}: {cnt} SAR detections (7d)",
            "region": label,
            "region_id": region_id,
            "sar_detections_7d": cnt,
            "total_entries": total,
            "category": "Dark Vessel Hotspot",
            "color": "#E040FB" if cnt > 0 else "#888",
        })
    return _payload("dark_vessels", items)


# ============================================================================
# Public registry — what to add to layers.LAYERS
# ============================================================================
EXTRA_LAYERS = [
    # 2026-05-23 Phase 2: tsunami/aurora/gdacs/submarine_cables/inaturalist
    # migrated to Vercel Cron (src/app/api/globe/<id>/cron/route.ts).
    # Weather expansion
    # {"id": "tsunami", "interval_s": 300, "fetch": fetch_tsunami},
    # {"id": "aurora", "interval_s": 1800, "fetch": fetch_aurora},
    # Multi-hazard + cyclones
    # {"id": "gdacs", "interval_s": 600, "fetch": fetch_gdacs},
    # Infra
    # {"id": "submarine_cables", "interval_s": 86400, "fetch": fetch_submarine_cables},
    {"id": "power_plants", "interval_s": 86400, "fetch": fetch_power_plants},
    {"id": "hospitals", "interval_s": 86400, "fetch": fetch_hospitals},
    {"id": "volcanoes", "interval_s": 86400, "fetch": fetch_volcanoes},
    {"id": "shipping_lanes", "interval_s": 86400, "fetch": fetch_shipping_lanes},
    # Maritime intel overlays
    {"id": "sanctioned_vessels", "interval_s": 600, "fetch": fetch_sanctioned_vessels},
    {"id": "port_congestion", "interval_s": 600, "fetch": fetch_port_congestion},
    # Ocean + atmosphere expanded
    {"id": "ndbc_buoys", "interval_s": 1800, "fetch": fetch_ndbc_buoys},
    {"id": "metar", "interval_s": 600, "fetch": fetch_metar},
    {"id": "tides", "interval_s": 86400, "fetch": fetch_tides},
    # Health + wildlife
    {"id": "who_outbreaks", "interval_s": 3600, "fetch": fetch_who_don},
    {"id": "wildlife_gbif", "interval_s": 3600, "fetch": fetch_gbif},
    # {"id": "inaturalist", "interval_s": 1800, "fetch": fetch_inaturalist},
    # Corporate / financial — SEC filings (keyless, all-US plotted at SEC HQ)
    {"id": "sec_filings", "interval_s": 3600, "fetch": fetch_sec_edgar},
    # Env-gated (free keys)
    {"id": "acled", "interval_s": 3600, "fetch": fetch_acled},
    {"id": "eia_grid", "interval_s": 1800, "fetch": fetch_eia_grid},
    {"id": "internet_outages", "interval_s": 1800, "fetch": fetch_internet_outages},
    {"id": "ebird", "interval_s": 3600, "fetch": fetch_ebird},
    {"id": "dark_vessels", "interval_s": 3600, "fetch": fetch_gfw_sar},
    # FRED — US macros (env-gated on FRED_API_KEY).
    {"id": "fred_macros", "interval_s": 3600, "fetch": fetch_fred_macros},
]
